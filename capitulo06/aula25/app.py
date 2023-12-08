from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / "workbook.xlsx"

workbook = Workbook()
sheet_name = "Planilha de notas"
workbook.create_sheet(sheet_name, 0)
worksheet: Worksheet = workbook[sheet_name]

workbook.remove(workbook["Sheet"])

worksheet.cell(1, 1, "Nome")
worksheet.cell(1, 2, "Idade")
worksheet.cell(1, 3, "Nota")

students = [
    ["João", 40, 5.5],
    ["Maria", 21, 7.0],
    ["Luiza", 23, 9.5],
    ["Alexandre", 30, 10],
]

# for i, student_row in enumerate(students,start=2):
#     for j, student_column in enumerate(student_row, start=1):
#         worksheet.cell(i, j, student_column)

for student in students:
    worksheet.append(student)

workbook.save(WORKBOOK_PATH)
